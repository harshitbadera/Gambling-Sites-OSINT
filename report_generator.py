"""
=============================================================================
GAMBLING OSINT - REPORT GENERATOR
=============================================================================
Generates comprehensive Excel reports with multiple sheets:
  - Discovery:       All discovered domains and their sources
  - Classification:  Classification scores and categories
  - Infrastructure:  Full DNS and hosting infrastructure data
  - Correlation:     Infrastructure cluster relationships
  - Summary Stats:   Aggregate statistics, top providers, distributions

Uses openpyxl for rich Excel formatting with headers, colors, and filters.
=============================================================================
"""

import os
import logging
from datetime import datetime
from collections import Counter

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

import config
from utils import read_csv_domains

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# STYLING CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

# Color palette
HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="2C3E6B", end_color="2C3E6B", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

# Classification colors
CATEGORY_COLORS = {
    "Gambling-Related": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    "Possibly Gambling-Related": PatternFill(start_color="FFA06B", end_color="FFA06B", fill_type="solid"),
    "Not Gambling-Related": PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),
    "India-Focused": PatternFill(start_color="FF8C42", end_color="FF8C42", fill_type="solid"),
    "Possibly India-Focused": PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
    "Unknown": PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"),
}

# Alternating row colors
ROW_FILL_EVEN = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
ROW_FILL_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Stats section colors
STAT_LABEL_FILL = PatternFill(start_color="E8EAF0", end_color="E8EAF0", fill_type="solid")
STAT_VALUE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)

DEFAULT_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1B2A4A")


# ─────────────────────────────────────────────────────────────────────────────
# REPORT GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

class ReportGenerator:
    """
    Generates a multi-sheet Excel report from all pipeline outputs.
    """

    def __init__(self, output_path: str = None):
        self.output_path = output_path or config.FINAL_REPORT_XLSX
        self.wb = openpyxl.Workbook()
        # Remove the default sheet
        self.wb.remove(self.wb.active)

    def generate(
        self,
        discovery_data: list = None,
        classification_data: list = None,
        enrichment_data: list = None,
        correlation_result: dict = None,
    ):
        """
        Generate the complete report.
        Loads data from CSV files if not provided directly.
        """
        logger.info("=" * 60)
        logger.info("GENERATING REPORT")
        logger.info("=" * 60)

        # Load data from CSV files if not provided
        if discovery_data is None and os.path.exists(config.DISCOVERY_CSV):
            discovery_data = read_csv_domains(config.DISCOVERY_CSV)
        if classification_data is None and os.path.exists(config.CLASSIFICATION_CSV):
            classification_data = read_csv_domains(config.CLASSIFICATION_CSV)
        if enrichment_data is None and os.path.exists(config.ENRICHMENT_CSV):
            enrichment_data = read_csv_domains(config.ENRICHMENT_CSV)
        if correlation_result is None and os.path.exists(config.CORRELATION_CSV):
            correlation_result = {"cluster_rows": read_csv_domains(config.CORRELATION_CSV)}

        # Generate sheets
        self._write_discovery_sheet(discovery_data or [])
        self._write_classification_sheet(classification_data or [])
        self._write_infrastructure_sheet(enrichment_data or [])
        self._write_correlation_sheet(correlation_result or {})
        self._write_summary_sheet(
            discovery_data or [],
            classification_data or [],
            enrichment_data or [],
            correlation_result or {},
        )

        # Save
        self.wb.save(self.output_path)
        logger.info(f"Report saved to: {self.output_path}")
        print(f"\n[REPORT] Saved to: {self.output_path}")

    # ── Helper methods ──────────────────────────────────────────────────

    def _style_header_row(self, ws, row: int, max_col: int):
        """Apply header styling to a row."""
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    def _style_data_rows(self, ws, start_row: int, end_row: int, max_col: int):
        """Apply alternating row colors and borders to data rows."""
        for row in range(start_row, end_row + 1):
            fill = ROW_FILL_EVEN if (row - start_row) % 2 == 0 else ROW_FILL_ODD
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = DEFAULT_FONT
                cell.fill = fill
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=False)

    def _auto_column_width(self, ws, max_col: int, max_rows: int = 100, max_width: int = 50):
        """Auto-adjust column widths based on content."""
        for col in range(1, max_col + 1):
            max_len = 0
            col_letter = get_column_letter(col)
            for row in range(1, min(max_rows + 1, ws.max_row + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            adjusted = min(max_len + 4, max_width)
            ws.column_dimensions[col_letter].width = max(adjusted, 12)

    def _add_autofilter(self, ws, max_col: int):
        """Add auto-filter to header row."""
        last_col = get_column_letter(max_col)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    def _write_data_rows(self, ws, headers: list, data: list) -> int:
        """Write headers and data rows. Returns the last row number."""
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        self._style_header_row(ws, 1, len(headers))

        # Write data
        for row_idx, record in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = record.get(header, "")
                ws.cell(row=row_idx, column=col_idx, value=value)

        last_row = len(data) + 1
        self._style_data_rows(ws, 2, last_row, len(headers))
        self._auto_column_width(ws, len(headers))
        self._add_autofilter(ws, len(headers))

        # Freeze top row
        ws.freeze_panes = "A2"

        return last_row

    # ── Sheet Writers ───────────────────────────────────────────────────

    def _write_discovery_sheet(self, data: list):
        """Write the Discovery sheet."""
        ws = self.wb.create_sheet(title=config.REPORT_SHEET_NAMES["discovery"])

        if not data:
            ws.cell(row=1, column=1, value="No discovery data available")
            return

        headers = ["domain", "source", "raw_input", "site_name", "discovered_at"]
        # Filter to only include headers that exist in the data
        available = set()
        for record in data:
            available.update(record.keys())
        headers = [h for h in headers if h in available]

        self._write_data_rows(ws, headers, data)
        logger.info(f"  Discovery sheet: {len(data)} rows")

    def _write_classification_sheet(self, data: list):
        """Write the Classification sheet with color-coded categories."""
        ws = self.wb.create_sheet(title=config.REPORT_SHEET_NAMES["classification"])

        if not data:
            ws.cell(row=1, column=1, value="No classification data available")
            return

        headers = [
            "domain", "source", "site_name", "fetch_status", "http_status_code",
            "page_title", "gambling_score", "india_score", "confidence",
            "classification", "india_classification", "combined_classification",
            "gambling_keywords_found", "india_keywords_found",
            "payment_methods_found", "fetch_error", "classified_at",
        ]
        available = set()
        for record in data:
            available.update(record.keys())
        headers = [h for h in headers if h in available]

        last_row = self._write_data_rows(ws, headers, data)

        # Color-code classification columns
        if "classification" in headers:
            class_col = headers.index("classification") + 1
            for row in range(2, last_row + 1):
                cell = ws.cell(row=row, column=class_col)
                category = str(cell.value or "")
                if category in CATEGORY_COLORS:
                    cell.fill = CATEGORY_COLORS[category]
                    cell.font = Font(name="Calibri", bold=True, size=10)

        if "india_classification" in headers:
            india_col = headers.index("india_classification") + 1
            for row in range(2, last_row + 1):
                cell = ws.cell(row=row, column=india_col)
                category = str(cell.value or "")
                if category in CATEGORY_COLORS:
                    cell.fill = CATEGORY_COLORS[category]
                    cell.font = Font(name="Calibri", bold=True, size=10)

        logger.info(f"  Classification sheet: {len(data)} rows")

    def _write_infrastructure_sheet(self, data: list):
        """Write the Infrastructure sheet."""
        ws = self.wb.create_sheet(title=config.REPORT_SHEET_NAMES["infrastructure"])

        if not data:
            ws.cell(row=1, column=1, value="No infrastructure data available")
            return

        headers = [
            "domain", "registrable_domain", "ipv4_addresses", "ipv6_addresses",
            "asn", "asn_description", "hosting_provider", "country",
            "nameservers", "mx_records", "txt_records", "soa_record",
            "network_cidr", "enrichment_status", "enrichment_error", "enriched_at",
        ]
        available = set()
        for record in data:
            available.update(record.keys())
        headers = [h for h in headers if h in available]

        self._write_data_rows(ws, headers, data)
        logger.info(f"  Infrastructure sheet: {len(data)} rows")

    def _write_correlation_sheet(self, correlation_result: dict):
        """Write the Correlation sheet showing infrastructure clusters."""
        ws = self.wb.create_sheet(title=config.REPORT_SHEET_NAMES["correlation"])

        # Check if we have cluster data
        clusters = correlation_result.get("clusters", {})
        cluster_rows = correlation_result.get("cluster_rows", [])

        if cluster_rows:
            # If loaded from CSV, write directly
            headers = ["domain", "cluster_dimension", "cluster_value", "cluster_size", "correlated_at"]
            available = set()
            for record in cluster_rows:
                available.update(record.keys())
            headers = [h for h in headers if h in available]
            self._write_data_rows(ws, headers, cluster_rows)
            logger.info(f"  Correlation sheet: {len(cluster_rows)} rows (from CSV)")
            return

        if not clusters:
            ws.cell(row=1, column=1, value="No correlation data available")
            return

        # Build rows from cluster data
        row_data = []
        for dimension, groups in clusters.items():
            for value, domains in sorted(groups.items(), key=lambda x: -len(x[1])):
                for domain in sorted(domains):
                    row_data.append({
                        "cluster_dimension": dimension,
                        "cluster_value": value,
                        "cluster_size": len(domains),
                        "domain": domain,
                    })

        headers = ["cluster_dimension", "cluster_value", "cluster_size", "domain"]
        self._write_data_rows(ws, headers, row_data)
        logger.info(f"  Correlation sheet: {len(row_data)} rows")

    def _write_summary_sheet(
        self,
        discovery_data: list,
        classification_data: list,
        enrichment_data: list,
        correlation_result: dict,
    ):
        """Write the Summary Statistics sheet."""
        ws = self.wb.create_sheet(title=config.REPORT_SHEET_NAMES["summary"])

        row = 1

        # ── Title ──
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        title_cell = ws.cell(row=row, column=1, value="GAMBLING OSINT — INTELLIGENCE REPORT")
        title_cell.font = TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center")
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(
            row=row, column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ).font = Font(name="Calibri", italic=True, size=10, color="666666")
        row += 2

        # ── Overview Stats ──
        row = self._write_stat_section(ws, row, "PIPELINE OVERVIEW", [
            ("Total Domains Discovered", len(discovery_data)),
            ("Total Domains Classified", len(classification_data)),
            ("Total Domains Enriched", len(enrichment_data)),
        ])
        row += 1

        # ── Classification Distribution ──
        if classification_data:
            class_counts = Counter(
                r.get("classification", "Unknown") for r in classification_data
            )
            row = self._write_stat_section(ws, row, "CLASSIFICATION DISTRIBUTION", [
                (cat, count)
                for cat, count in class_counts.most_common()
            ])
            row += 1

            india_counts = Counter(
                r.get("india_classification", "") for r in classification_data
                if r.get("india_classification")
            )
            if india_counts:
                row = self._write_stat_section(ws, row, "INDIA-TARGETING DISTRIBUTION", [
                    (cat, count)
                    for cat, count in india_counts.most_common()
                ])
                row += 1

        # ── Top Hosting Providers ──
        if enrichment_data:
            provider_counts = Counter(
                r.get("hosting_provider", "Unknown") or "Unknown"
                for r in enrichment_data
            )
            row = self._write_stat_section(
                ws, row,
                f"TOP {config.SUMMARY_TOP_N} HOSTING PROVIDERS",
                list(provider_counts.most_common(config.SUMMARY_TOP_N)),
            )
            row += 1

            # ── Top ASNs ──
            asn_counts = Counter(
                f"{r.get('asn', 'N/A')} — {r.get('asn_description', '')}"
                for r in enrichment_data
                if r.get("asn")
            )
            row = self._write_stat_section(
                ws, row,
                f"TOP {config.SUMMARY_TOP_N} ASNs",
                list(asn_counts.most_common(config.SUMMARY_TOP_N)),
            )
            row += 1

            # ── Top Countries ──
            country_counts = Counter(
                r.get("country", "Unknown") or "Unknown"
                for r in enrichment_data
            )
            row = self._write_stat_section(
                ws, row,
                f"TOP {config.SUMMARY_TOP_N} COUNTRIES",
                list(country_counts.most_common(config.SUMMARY_TOP_N)),
            )
            row += 1

            # ── Top Nameservers ──
            ns_counter = Counter()
            for r in enrichment_data:
                ns_raw = r.get("nameservers", "")
                if ns_raw:
                    for ns in ns_raw.split("; "):
                        ns = ns.strip().rstrip(".")
                        if ns:
                            ns_counter[ns] += 1

            if ns_counter:
                row = self._write_stat_section(
                    ws, row,
                    f"TOP {config.SUMMARY_TOP_N} NAMESERVERS",
                    list(ns_counter.most_common(config.SUMMARY_TOP_N)),
                )
                row += 1

        # ── Discovery Sources ──
        if discovery_data:
            source_counts = Counter(
                r.get("source", "Unknown") for r in discovery_data
            )
            row = self._write_stat_section(ws, row, "DISCOVERY SOURCES", [
                (src, count) for src, count in source_counts.most_common()
            ])

        # Auto-width
        self._auto_column_width(ws, 4, max_rows=row)

        logger.info(f"  Summary sheet: {row} rows")

    def _write_stat_section(
        self, ws, start_row: int, title: str, items: list
    ) -> int:
        """
        Write a statistics section with a title and label-value pairs.
        Returns the next available row.
        """
        # Section header
        ws.merge_cells(
            start_row=start_row, start_column=1,
            end_row=start_row, end_column=4,
        )
        header_cell = ws.cell(row=start_row, column=1, value=title)
        header_cell.font = SUBHEADER_FONT
        header_cell.fill = SUBHEADER_FILL
        header_cell.alignment = Alignment(horizontal="left")
        for col in range(1, 5):
            ws.cell(row=start_row, column=col).fill = SUBHEADER_FILL
            ws.cell(row=start_row, column=col).border = THIN_BORDER
        start_row += 1

        # Column headers
        ws.cell(row=start_row, column=1, value="Item").font = BOLD_FONT
        ws.cell(row=start_row, column=2, value="Count").font = BOLD_FONT
        ws.cell(row=start_row, column=1).fill = STAT_LABEL_FILL
        ws.cell(row=start_row, column=2).fill = STAT_LABEL_FILL
        ws.cell(row=start_row, column=1).border = THIN_BORDER
        ws.cell(row=start_row, column=2).border = THIN_BORDER
        start_row += 1

        # Data rows
        for label, value in items:
            ws.cell(row=start_row, column=1, value=str(label)).font = DEFAULT_FONT
            ws.cell(row=start_row, column=1).border = THIN_BORDER
            val_cell = ws.cell(row=start_row, column=2, value=value)
            val_cell.font = BOLD_FONT
            val_cell.alignment = Alignment(horizontal="center")
            val_cell.border = THIN_BORDER
            start_row += 1

        return start_row


# ─────────────────────────────────────────────────────────────────────────────
# CONVENIENCE FUNCTION
# ─────────────────────────────────────────────────────────────────────────────

def generate_report(
    discovery_data: list = None,
    classification_data: list = None,
    enrichment_data: list = None,
    correlation_result: dict = None,
    output_path: str = None,
):
    """
    Generate the full Excel report.
    Loads from CSV files if data is not provided directly.
    """
    generator = ReportGenerator(output_path=output_path)
    generator.generate(
        discovery_data=discovery_data,
        classification_data=classification_data,
        enrichment_data=enrichment_data,
        correlation_result=correlation_result,
    )
    return generator.output_path


# ─────────────────────────────────────────────────────────────────────────────
# STANDALONE EXECUTION
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    config.setup_logging()
    logger.info("Running Report Generator standalone")

    report_path = generate_report()
    print(f"\n[REPORT] Generated: {report_path}")
